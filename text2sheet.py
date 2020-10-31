#! python
# script responsible for transer all .txt files from cwd to newly created .xlsx file
# X 2020 Arnold Cytrowski

import os, sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# opcje:
# sys.argv == 2: searching for all textfiles in cwd
# sys.argv > 2 user have to type all the files which he wants to copy into the worksheet

choice = input('hello there. Instructions:\n\t0 - press 0 to search for all .txt files in the cwd\nor\n\ttype all the files which you want to put in the new workbook, separating them by space\n')


cwd = os.getcwd()
text_files = []
if choice == '0':
    for filename in os.listdir(cwd):
        if filename.endswith('.txt'):
            text_files.append(os.path.join(cwd, filename))
    
else:
    filenames = choice.split()
    for filename in filenames:
        if filename.endswith('.txt'):
            text_files.append(os.path.join(cwd, filename))

wb = openpyxl.Workbook()
sheet = wb.active

col_num = 1



for file in text_files:


    row_num = 2
    longest_col = 0

    bold = Font(bold=True)
    sheet.cell(row=1, column=col_num).value = os.path.basename(file)
    sheet.cell(row=1, column=col_num).font = bold

    lines = open(file, encoding='UTF-8').readlines()

    for line in lines:
        line = line.strip()

        if len(line) > longest_col:
            longest_col = len(line)

        sheet.cell(row = row_num, column = col_num).value = line
        row_num += 1

    column_letter = get_column_letter(col_num)
    sheet.column_dimensions[column_letter].width = longest_col + 5
    col_num += 1

wb.save(os.path.join(cwd, 'text2sheet.xlsx'))

print('aaand it\'s done, bye bye')

        








